#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
增强版代码统计工具
- 支持 C++ / CUDA / QML / UI 多维统计
- 输出终端汇总、TopN 文件
- 可导出 JSON
- 可生成图表（需要 matplotlib）
"""

import argparse
import json
import os
from collections import defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple


EXCLUDE_DIRS = {
    ".git",
    ".svn",
    ".hg",
    "CVS",
    "cmake-build-debug",
    "cmake-build-release",
    "build",
    "dist",
    "venv",
    ".venv",
    "__pycache__",
    "node_modules",
    "bin",
    "obj",
    "Debug",
    "Release",
    "x64",
    "x86",
    ".idea",
    ".vscode",
    "QXlsx",
    "OpenCV",
    "include",
    "lib",
    "exprtk",
    "release",
}

EXCLUDE_FILES = {
    "__init__.py",
    ".gitignore",
    ".gitattributes",
    "README.md",
    "README.txt",
    "FUTURE.md",
    "LICENSE",
    "LICENSE.txt",
    "NOTICE",
    "NOTICE.txt",
}

FILE_TYPE_GROUPS = {
    "C++": [".h", ".hpp", ".cpp", ".c", ".cc"],
    "QML": [".qml"],
    "UI": [".ui"],
    "CUDA": [".cuh", ".cu"],
}

C_LIKE_EXTS = {".h", ".hpp", ".cpp", ".c", ".cc", ".cu", ".cuh", ".qml"}
XML_LIKE_EXTS = {".ui"}


@dataclass
class LineStats:
    total: int = 0
    code: int = 0
    comment: int = 0
    blank: int = 0

    def add(self, other: "LineStats") -> None:
        self.total += other.total
        self.code += other.code
        self.comment += other.comment
        self.blank += other.blank

    @property
    def code_ratio(self) -> float:
        return (self.code / self.total * 100.0) if self.total else 0.0


@dataclass
class FileResult:
    path: str
    group: str
    ext: str
    stats: LineStats


def get_file_group(ext: str) -> Optional[str]:
    for group, exts in FILE_TYPE_GROUPS.items():
        if ext in exts:
            return group
    return None


def analyze_c_like(lines: Iterable[str]) -> LineStats:
    stats = LineStats()
    in_block_comment = False

    for raw in lines:
        stats.total += 1
        line = raw.rstrip("\n")

        if not line.strip():
            stats.blank += 1
            continue

        i = 0
        has_code = False
        has_comment = False

        while i < len(line):
            if in_block_comment:
                has_comment = True
                end_idx = line.find("*/", i)
                if end_idx == -1:
                    i = len(line)
                else:
                    in_block_comment = False
                    i = end_idx + 2
                continue

            if line.startswith("//", i):
                has_comment = True
                break

            if line.startswith("/*", i):
                has_comment = True
                in_block_comment = True
                i += 2
                continue

            if not line[i].isspace():
                has_code = True
            i += 1

        if has_code:
            stats.code += 1
        elif has_comment:
            stats.comment += 1
        else:
            # 理论兜底，保证 total = code + comment + blank
            stats.blank += 1

    return stats


def analyze_xml_like(lines: Iterable[str]) -> LineStats:
    stats = LineStats()
    in_block_comment = False

    for raw in lines:
        stats.total += 1
        line = raw.rstrip("\n")

        if not line.strip():
            stats.blank += 1
            continue

        i = 0
        has_code = False
        has_comment = False

        while i < len(line):
            if in_block_comment:
                has_comment = True
                end_idx = line.find("-->", i)
                if end_idx == -1:
                    i = len(line)
                else:
                    in_block_comment = False
                    i = end_idx + 3
                continue

            if line.startswith("<!--", i):
                has_comment = True
                in_block_comment = True
                i += 4
                continue

            if not line[i].isspace():
                has_code = True
            i += 1

        if has_code:
            stats.code += 1
        elif has_comment:
            stats.comment += 1
        else:
            stats.blank += 1

    return stats


def count_file(file_path: Path) -> Optional[FileResult]:
    ext = file_path.suffix.lower()
    group = get_file_group(ext)
    if group is None:
        return None

    try:
        lines = file_path.read_text(encoding="utf-8", errors="ignore").splitlines(keepends=True)
    except Exception as exc:
        print(f"[WARN] 读取失败: {file_path} ({exc})")
        return None

    if ext in C_LIKE_EXTS:
        stats = analyze_c_like(lines)
    elif ext in XML_LIKE_EXTS:
        stats = analyze_xml_like(lines)
    else:
        return None

    return FileResult(path=str(file_path), group=group, ext=ext, stats=stats)


def format_table(rows: List[List[str]], headers: List[str]) -> str:
    all_rows = [headers] + rows
    widths = [max(len(str(row[i])) for row in all_rows) for i in range(len(headers))]

    def fmt(row: List[str]) -> str:
        return "  ".join(str(value).ljust(widths[i]) for i, value in enumerate(row))

    sep = "  ".join("-" * width for width in widths)
    lines = [fmt(headers), sep]
    lines.extend(fmt(row) for row in rows)
    return "\n".join(lines)


def safe_relpath(path: str, root: Path) -> str:
    try:
        return str(Path(path).resolve().relative_to(root.resolve()))
    except Exception:
        return path


def text_bar_chart(data: List[Tuple[str, int]], width: int = 40) -> str:
    if not data:
        return "(无数据)"
    max_value = max(v for _, v in data)
    if max_value <= 0:
        return "\n".join(f"{name:20} |" for name, _ in data)
    lines = []
    for name, value in data:
        bar_len = int(value / max_value * width)
        bar = "#" * bar_len
        lines.append(f"{name[:20]:20} | {bar} {value}")
    return "\n".join(lines)


def generate_plots(
    output_dir: Path,
    group_stats: Dict[str, LineStats],
    top_files: List[FileResult],
    total_stats: LineStats,
    root_dir: Path,
) -> bool:
    try:
        import matplotlib.pyplot as plt
    except Exception:
        print("[WARN] 未安装 matplotlib，跳过 PNG 图表生成。")
        print("       可执行: pip install matplotlib")
        return False

    output_dir.mkdir(parents=True, exist_ok=True)

    # 图1：分组代码行数
    groups = [g for g, s in group_stats.items() if s.total > 0]
    code_values = [group_stats[g].code for g in groups]
    if groups:
        plt.figure(figsize=(10, 5))
        plt.bar(groups, code_values)
        plt.title("Code Lines by Group")
        plt.ylabel("Code Lines")
        plt.tight_layout()
        plt.savefig(output_dir / "group_code_lines.png", dpi=160)
        plt.close()

    # 图2：TopN 文件代码行
    if top_files:
        labels = [safe_relpath(f.path, root_dir) for f in top_files]
        values = [f.stats.code for f in top_files]
        plt.figure(figsize=(12, 6))
        plt.barh(labels[::-1], values[::-1])
        plt.title("Top Files by Code Lines")
        plt.xlabel("Code Lines")
        plt.tight_layout()
        plt.savefig(output_dir / "top_files_code_lines.png", dpi=160)
        plt.close()

    # 图3：代码/注释/空行占比
    pie_labels = ["Code", "Comment", "Blank"]
    pie_values = [total_stats.code, total_stats.comment, total_stats.blank]
    if sum(pie_values) > 0:
        plt.figure(figsize=(6, 6))
        plt.pie(pie_values, labels=pie_labels, autopct="%1.1f%%", startangle=90)
        plt.title("Line Composition")
        plt.tight_layout()
        plt.savefig(output_dir / "line_composition.png", dpi=160)
        plt.close()

    print(f"[INFO] 图表已生成到: {output_dir}")
    return True


def export_excel_report(
    excel_path: Path,
    root_dir: Path,
    total_stats: LineStats,
    group_stats: Dict[str, LineStats],
    dir_stats: Dict[str, LineStats],
    files_data: List[FileResult],
) -> bool:
    try:
        from openpyxl import Workbook
    except Exception:
        print("[WARN] 未安装 openpyxl，无法导出 Excel。")
        print("       可执行: pip install openpyxl")
        return False

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Root", str(root_dir.resolve())])
    ws_summary.append(["GeneratedAt", datetime.now().isoformat(timespec="seconds")])
    ws_summary.append([])
    ws_summary.append(["Total", "Code", "Comment", "Blank", "CodeRatio(%)"])
    ws_summary.append(
        [
            total_stats.total,
            total_stats.code,
            total_stats.comment,
            total_stats.blank,
            round(total_stats.code_ratio, 2),
        ]
    )

    ws_groups = wb.create_sheet("Groups")
    ws_groups.append(["Group", "Total", "Code", "Comment", "Blank", "CodeRatio(%)"])
    for group, stats in sorted(group_stats.items(), key=lambda x: x[1].code, reverse=True):
        ws_groups.append([group, stats.total, stats.code, stats.comment, stats.blank, round(stats.code_ratio, 2)])

    ws_dirs = wb.create_sheet("Directories")
    ws_dirs.append(["Directory", "Total", "Code", "Comment", "Blank", "CodeRatio(%)"])
    for directory, stats in sorted(dir_stats.items(), key=lambda x: x[1].code, reverse=True):
        ws_dirs.append([directory, stats.total, stats.code, stats.comment, stats.blank, round(stats.code_ratio, 2)])

    ws_files = wb.create_sheet("Files")
    ws_files.append(["Path", "Group", "Ext", "Total", "Code", "Comment", "Blank", "CodeRatio(%)"])
    for item in files_data:
        ws_files.append(
            [
                safe_relpath(item.path, root_dir),
                item.group,
                item.ext,
                item.stats.total,
                item.stats.code,
                item.stats.comment,
                item.stats.blank,
                round(item.stats.code_ratio, 2),
            ]
        )

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    print(f"[INFO] Excel 已导出: {excel_path}")
    return True


def update_trend_history(
    trend_store: Path,
    root_dir: Path,
    total_stats: LineStats,
    file_count: int,
    group_stats: Dict[str, LineStats],
) -> Tuple[dict, Optional[dict], List[dict]]:
    snapshot = {
        "timestamp": datetime.now().isoformat(timespec="seconds"),
        "root": str(root_dir.resolve()),
        "file_count": file_count,
        "total": asdict(total_stats),
        "groups": {k: asdict(v) for k, v in group_stats.items()},
    }

    history: List[dict] = []
    if trend_store.exists():
        try:
            lines = trend_store.read_text(encoding="utf-8").splitlines()
            history = [json.loads(line) for line in lines if line.strip()]
        except Exception as exc:
            print(f"[WARN] 读取趋势历史失败，改为重建历史: {exc}")
            history = []

    prev = history[-1] if history else None
    history.append(snapshot)

    trend_store.parent.mkdir(parents=True, exist_ok=True)
    trend_store.write_text("\n".join(json.dumps(item, ensure_ascii=False) for item in history) + "\n", encoding="utf-8")
    print(f"[INFO] 趋势历史已更新: {trend_store}")
    return snapshot, prev, history


def print_trend_delta(snapshot: dict, prev: Optional[dict]) -> None:
    print("\n[趋势统计]")
    if prev is None:
        print("首次记录，暂无可比较的上一快照。")
        return

    curr_total = snapshot["total"]
    prev_total = prev.get("total", {})
    curr_files = snapshot.get("file_count", 0)
    prev_files = prev.get("file_count", 0)

    def delta(name: str) -> str:
        diff = int(curr_total.get(name, 0)) - int(prev_total.get(name, 0))
        sign = "+" if diff >= 0 else ""
        return f"{sign}{diff}"

    rows = [
        ["文件数", str(curr_files), f"{curr_files - prev_files:+d}"],
        ["总行", str(curr_total.get("total", 0)), delta("total")],
        ["代码行", str(curr_total.get("code", 0)), delta("code")],
        ["注释行", str(curr_total.get("comment", 0)), delta("comment")],
        ["空行", str(curr_total.get("blank", 0)), delta("blank")],
    ]
    print(format_table(rows, ["指标", "当前值", "较上次变化"]))


def generate_trend_plot(history: List[dict], trend_limit: int, output_dir: Path) -> bool:
    try:
        import matplotlib.pyplot as plt
    except Exception:
        print("[WARN] 未安装 matplotlib，跳过趋势图生成。")
        return False

    data = history[-max(1, trend_limit):]
    if not data:
        print("[WARN] 趋势历史为空，无法生成趋势图。")
        return False

    labels = [item.get("timestamp", "") for item in data]
    code_lines = [int(item.get("total", {}).get("code", 0)) for item in data]
    comment_lines = [int(item.get("total", {}).get("comment", 0)) for item in data]
    blank_lines = [int(item.get("total", {}).get("blank", 0)) for item in data]

    output_dir.mkdir(parents=True, exist_ok=True)
    plt.figure(figsize=(12, 5))
    plt.plot(labels, code_lines, marker="o", label="Code")
    plt.plot(labels, comment_lines, marker="o", label="Comment")
    plt.plot(labels, blank_lines, marker="o", label="Blank")
    plt.title("Codebase Trend")
    plt.xlabel("Snapshot Time")
    plt.ylabel("Lines")
    plt.xticks(rotation=25, ha="right")
    plt.legend()
    plt.tight_layout()
    output_path = output_dir / "trend_lines.png"
    plt.savefig(output_path, dpi=160)
    plt.close()
    print(f"[INFO] 趋势图已生成: {output_path}")
    return True


def run_analysis(
    root_dir: Path,
    include_groups: Optional[List[str]],
    extra_exclude_dirs: List[str],
    sort_by: str,
    top_n: int,
) -> Tuple[List[FileResult], List[FileResult], Dict[str, LineStats], Dict[str, LineStats], LineStats]:
    files_data: List[FileResult] = []

    exclude_dirs = set(EXCLUDE_DIRS)
    exclude_dirs.update(extra_exclude_dirs)

    allowed_groups = set(FILE_TYPE_GROUPS.keys())
    if include_groups:
        allowed_groups = {g for g in include_groups if g in FILE_TYPE_GROUPS}

    for root, dirs, files in os.walk(root_dir):
        dirs[:] = [d for d in dirs if d not in exclude_dirs]

        for file_name in files:
            if file_name in EXCLUDE_FILES:
                continue
            file_path = Path(root) / file_name
            result = count_file(file_path)
            if result is None or result.group not in allowed_groups:
                continue
            files_data.append(result)

    group_stats = defaultdict(LineStats)
    dir_stats = defaultdict(LineStats)
    total_stats = LineStats()

    for item in files_data:
        group_stats[item.group].add(item.stats)
        rel_dir = str(Path(item.path).resolve().parent.relative_to(root_dir.resolve()))
        dir_stats[rel_dir if rel_dir else "."].add(item.stats)
        total_stats.add(item.stats)

    key_map = {
        "code": lambda x: x.stats.code,
        "total": lambda x: x.stats.total,
        "comment": lambda x: x.stats.comment,
        "ratio": lambda x: x.stats.code_ratio,
    }
    files_data.sort(key=key_map[sort_by], reverse=True)
    top_files = files_data[:top_n]

    return files_data, top_files, dict(group_stats), dict(dir_stats), total_stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="增强版 C++/CUDA/QML/UI 代码统计工具")
    parser.add_argument("directory", nargs="?", default=".", help="要统计的目录路径，默认当前目录")
    parser.add_argument("--top", type=int, default=15, help="展示 Top N 文件（默认 15）")
    parser.add_argument(
        "--sort-by",
        choices=["code", "total", "comment", "ratio"],
        default="code",
        help="Top 文件排序字段",
    )
    parser.add_argument(
        "--include",
        nargs="*",
        default=None,
        help="只统计指定分组，可选值: C++ QML UI CUDA",
    )
    parser.add_argument(
        "--exclude-dir",
        nargs="*",
        default=[],
        help="额外排除目录名（可多个）",
    )
    parser.add_argument("--json", dest="json_path", default=None, help="导出 JSON 文件路径")
    parser.add_argument("--excel", dest="excel_path", default=None, help="导出 Excel 文件路径（xlsx）")
    parser.add_argument("--plot", action="store_true", help="生成图表 PNG（需 matplotlib）")
    parser.add_argument(
        "--output-dir",
        default="stats_output",
        help="图表与输出文件目录（默认 stats_output）",
    )
    parser.add_argument("--trend", action="store_true", help="记录并展示趋势统计（与上次快照比较）")
    parser.add_argument(
        "--trend-store",
        default="stats_output/trend_history.jsonl",
        help="趋势历史存储路径（jsonl）",
    )
    parser.add_argument("--trend-plot", action="store_true", help="生成趋势折线图（需 matplotlib）")
    parser.add_argument("--trend-limit", type=int, default=20, help="趋势图最多使用最近 N 条快照")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    root_dir = Path(args.directory)
    if not root_dir.is_dir():
        print(f"错误：{root_dir} 不是有效目录")
        return

    files_data, top_files, group_stats, dir_stats, total_stats = run_analysis(
        root_dir=root_dir,
        include_groups=args.include,
        extra_exclude_dirs=args.exclude_dir,
        sort_by=args.sort_by,
        top_n=max(1, args.top),
    )

    print(f"统计目录: {root_dir.resolve()}")
    print("=" * 72)

    group_rows = []
    for group, stats in sorted(group_stats.items(), key=lambda x: x[1].code, reverse=True):
        group_rows.append(
            [
                group,
                str(stats.total),
                str(stats.code),
                str(stats.comment),
                str(stats.blank),
                f"{stats.code_ratio:.2f}%",
            ]
        )

    if group_rows:
        print("[分组统计]")
        print(
            format_table(
                group_rows,
                ["分组", "总行", "代码行", "注释行", "空行", "代码占比"],
            )
        )
    else:
        print("[分组统计] 无匹配文件")

    print("\n[总计]")
    total_row = [[str(total_stats.total), str(total_stats.code), str(total_stats.comment), str(total_stats.blank), f"{total_stats.code_ratio:.2f}%"]]
    print(format_table(total_row, ["总行", "代码行", "注释行", "空行", "代码占比"]))

    print("\n[Top 文件]")
    top_rows = []
    for item in top_files:
        top_rows.append(
            [
                safe_relpath(item.path, root_dir),
                item.group,
                str(item.stats.total),
                str(item.stats.code),
                str(item.stats.comment),
                str(item.stats.blank),
                f"{item.stats.code_ratio:.2f}%",
            ]
        )
    if top_rows:
        print(
            format_table(
                top_rows,
                ["文件", "分组", "总行", "代码行", "注释行", "空行", "代码占比"],
            )
        )
    else:
        print("无匹配文件")

    print("\n[分组代码量文本图]")
    bar_data = sorted(((k, v.code) for k, v in group_stats.items()), key=lambda x: x[1], reverse=True)
    print(text_bar_chart(bar_data))

    dir_top = sorted(dir_stats.items(), key=lambda x: x[1].code, reverse=True)[:10]
    print("\n[目录 Top10 代码量]")
    if dir_top:
        dir_rows = [[k, str(v.code), str(v.total), f"{v.code_ratio:.2f}%"] for k, v in dir_top]
        print(format_table(dir_rows, ["目录", "代码行", "总行", "代码占比"]))
    else:
        print("无匹配目录")

    output_dir = Path(args.output_dir)
    if args.plot:
        generate_plots(output_dir, group_stats, top_files, total_stats, root_dir)

    if args.excel_path:
        export_excel_report(
            excel_path=Path(args.excel_path),
            root_dir=root_dir,
            total_stats=total_stats,
            group_stats=group_stats,
            dir_stats=dir_stats,
            files_data=files_data,
        )

    if args.json_path:
        payload = {
            "root": str(root_dir.resolve()),
            "summary": asdict(total_stats),
            "groups": {k: asdict(v) for k, v in group_stats.items()},
            "dirs": {k: asdict(v) for k, v in dir_stats.items()},
            "top_files": [
                {
                    "path": safe_relpath(item.path, root_dir),
                    "group": item.group,
                    "ext": item.ext,
                    "stats": asdict(item.stats),
                }
                for item in top_files
            ],
        }
        out_path = Path(args.json_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"[INFO] JSON 已导出: {out_path}")

    if args.trend or args.trend_plot:
        snapshot, prev, history = update_trend_history(
            trend_store=Path(args.trend_store),
            root_dir=root_dir,
            total_stats=total_stats,
            file_count=len(files_data),
            group_stats=group_stats,
        )
        if args.trend:
            print_trend_delta(snapshot, prev)
        if args.trend_plot:
            generate_trend_plot(history, max(1, args.trend_limit), output_dir)


if __name__ == "__main__":
    main()